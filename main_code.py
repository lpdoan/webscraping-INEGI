### This is the script to download data from INEGI MEXICO (MX) ###
# The government data that we need are Precios al Consumidor (INPC) and INPP mercancías excluyendo petróleo (INPP)
# Latest Update: July 2023
# import packages
from bs4 import BeautifulSoup, Tag
import requests
import time
import pandas as pd
from selenium import webdriver # selenium works with python 3.7 and above
from selenium.webdriver.common.by import By
import openpyxl
from openpyxl import load_workbook # export and save data to "output" excel file
import dateparser # parse spanish months

# time tracking
start_time = time.time()

# open an Edge browser and load the page at the given URL
driver = webdriver.Edge()
INPC_INPP_URL = "https://www.inegi.org.mx/app/indicesdeprecios/Estructura.aspx?idEstructura=112001300020&T=%C3%8Dndices+de+Precios+al+Consumidor&ST=Principales+%C3%ADndices+(mensual)?enablejsapi=1"
driver.get(INPC_INPP_URL)
time.sleep(10)

# send a get request a second time to load all important elements on the webpage
driver.get(INPC_INPP_URL)
date_element = driver.find_element(By.ID, "MainContent_wuc_Arbol1_theadTable")
inpc_element = driver.find_element(By.ID, "TR_11200130002000200010")
inpp_element = driver.find_element(By.ID, "TR_11200130002000200030")

time.sleep(1)

# extract the inpc and inpp data by splitting the inpc_element and inpp_element strings using the newline "\n" delimiter
inpc= inpc_element.text.split("\n") # inpc is a list including the series' name and 3 months' data e.g., ['Precios al Consumidor (INPC)', '128.363', '128.084', '128.214']
inpp = inpp_element.text.split("\n") # inpp is a list including the series' name and 3 months' data e.g., ['INPP mercancías excluyendo petróleo', '124.017', '123.497', '122.968']

# the dates (month-year format) corresponding to the data
inegi_dates = date_element.find_elements(By.TAG_NAME, "th")
inegi_dates = [element.text for element in inegi_dates[1:4]]

# Convert date time information from string to datetime
inegi_dates = [dateparser.parse(inegi_dates[i]) for i in range(len(inegi_dates))]

# create the final dataframe
# df has the series names as column names and inegi dates 
df = pd.DataFrame([inpc[1:], inpp[1:]]).T
df.index = inegi_dates
df.columns = [inpc[0], inpp[0]]

# Export to excel
output_name = "webscraping_INEGI_" + dt.now().strftime("%d-%m-%Y") + ".xlsx"
wb = openpyxl.Workbook()
inegi_ws = wb.active

# store the output in a folder named "output" in the directory
path = "./output/" + output_name
wb.save(path)
inegi_obj = openpyxl.load_workbook(path)
writer = pd.ExcelWriter(path, engine = "openpyxl")
df.to_excel(writer, sheet_name = "INEGI_INPP_INPC_MEXICO")
writer.close()

# Print runtime of whole script
print("The script completed the data retrieval in %s seconds." %
    (time.time() - start_time))
